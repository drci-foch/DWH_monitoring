import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from queries import DatabaseQualityChecker
from datetime import datetime
import logging
import time
import threading
import sys

def configure_logging():
    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        filename='database_quality_report.log',
        filemode='w'
    )
    # Also output to console
    console = logging.StreamHandler()
    console.setLevel(logging.INFO)
    formatter = logging.Formatter('%(name)-12s: %(levelname)-8s %(message)s')
    console.setFormatter(formatter)
    logging.getLogger('').addHandler(console)


class LoadingIndicator:
    def __init__(self, description="Loading"):
        self.description = description
        self.is_running = False
        self.animation = "|/-\\"
        self.idx = 0
        self.thread = None

    def animate(self):
        while self.is_running:
            print(f"\r{self.description} {self.animation[self.idx % len(self.animation)]}", end="")
            self.idx += 1
            time.sleep(0.1)

    def start(self):
        self.is_running = True
        self.thread = threading.Thread(target=self.animate)
        self.thread.start()

    def stop(self):
        self.is_running = False
        if self.thread:
            self.thread.join()
        sys.stdout.write('\r' + ' ' * (len(self.description) + 2) + '\r')
        sys.stdout.flush()

logger = logging.getLogger(__name__)

class DatabaseQualityReportGenerator:
    def __init__(self):
        self.db_checker = DatabaseQualityChecker()
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
        self.logger = logging.getLogger(__name__)

    def generate_report(self, filename=r"P:\RechercheClinique\Unité DATA\EDS Foch\Documentation EDS\Documentation technique\Documentation_base_DWH\Monitoring DWH\database_quality_report_{}.xlsx"):
        current_date = datetime.now().strftime("%d-%m-%Y")
        filename = filename.format(current_date)
        
        self.logger.info("Starting report generation...")
        print("Starting report generation...")
        
        try:
            self.all_stats = self.db_checker.get_all_statistics()
            self.logger.info("Data fetched successfully.")
            print("Data fetched successfully.")
            
            self.logger.info("Generating report sheets...")
            print("Generating report sheets...")
            self.add_summary_sheet()
            self.add_document_metrics_sheet()
            self.add_document_counts_sheet()
            self.add_recent_document_counts_sheet()
            self.add_top_users_sheet()
            self.add_top_users_current_year_sheet()
            self.add_archive_status_sheet()
            
            self.workbook.save(filename)
            self.logger.info(f"Report generated: {filename}")
            print(f"Report generated: {filename}")
        except Exception as e:
            self.logger.exception("An error occurred during report generation:")
            print(f"An error occurred: {str(e)}")


    def add_archive_status_sheet(self):
        sheet = self.workbook.create_sheet("Archive Status")
        sheet.column_dimensions['A'].width = 40
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20

        sheet['A1'] = "Archive Status Report"
        sheet['A1'].font = Font(bold=True, size=14)
        sheet.merge_cells('A1:C1')

        # Add archive period information
        archive_period = self.all_stats['archive_period']
        sheet['A3'] = "Current Archive Period (years)"
        sheet['B3'] = archive_period
        sheet['B3'].number_format = '0.00'

        if archive_period > 20:
            sheet['B3'].font = Font(color="FF0000", bold=True)
            sheet['C3'] = "Exceeds 20-year limit"
            sheet['C3'].font = Font(color="FF0000", italic=True)

        # Add total documents to suppress
        total_to_suppress = self.all_stats['total_documents_to_suppress'][0][0]
        sheet['A5'] = "Total Documents to Suppress"
        sheet['B5'] = total_to_suppress
        sheet['B5'].number_format = '#,##0'

        # Add table headers
        sheet['A7'] = "Document Origin Code"
        sheet['B7'] = "Documents to Suppress"
        sheet['C7'] = "Percentage of Total"

        # Get documents to suppress by origin
        documents_to_suppress = self.all_stats['documents_to_suppress']

        # Add data to the table
        for i, (origin, count) in enumerate(documents_to_suppress, start=8):
            sheet[f'A{i}'] = origin
            sheet[f'B{i}'] = count
            sheet[f'B{i}'].number_format = '#,##0'
            sheet[f'C{i}'] = f'=B{i}/B5'
            sheet[f'C{i}'].number_format = '0.00%'

        # Apply styling
        self.apply_table_style(sheet, f'A7:C{7+len(documents_to_suppress)}')

        # Add a pie chart
        self.add_pie_chart(sheet, f'A7:C{7+len(documents_to_suppress)}', 'E7', "Documents to Suppress by Origin")

    def add_document_count_graph(self, sheet, origin_code, by_year=True):
        if by_year:
            counts = self.all_stats['document_counts_by_year'].get(origin_code, [])
            title = f"Document Count by Year - {origin_code}"
            x_axis_title = "Year"
        else:
            counts = self.all_stats['recent_document_counts_by_month'].get(origin_code, [])
            title = f"Recent Document Count by Month - {origin_code}"
            x_axis_title = "Month"

        if not counts:
            self.logger.warning(f"No data available for {origin_code}")
            return

        # Add data to sheet
        start_row = sheet.max_row + 2
        sheet.cell(row=start_row, column=1, value=title)
        sheet.cell(row=start_row + 1, column=1, value=x_axis_title)
        sheet.cell(row=start_row + 1, column=2, value="Document Count")

        for i, (date, count) in enumerate(counts, start=start_row + 2):
            sheet.cell(row=i, column=1, value=date)
            sheet.cell(row=i, column=2, value=count)

        # Create chart
        chart = LineChart()
        chart.title = title
        chart.style = 2
        chart.x_axis.title = x_axis_title
        chart.y_axis.title = "Document Count"

        # Add data series
        data = Reference(sheet, min_col=2, min_row=start_row + 1, max_row=start_row + 1 + len(counts))
        series = chart.series.new()
        series.values = data
        series.title = "Document Count"

        # Set categories
        cats = Reference(sheet, min_col=1, min_row=start_row + 2, max_row=start_row + 1 + len(counts))
        chart.set_categories(cats)

        # Add median line
        median = self.db_checker.get_median_document_count(counts)
        median_series = chart.series.new()
        median_series.values = [median] * len(counts)
        median_series.title = f"Median ({median})"

        # Style the series
        chart.series[0].graphicalProperties.line.width = 20000  # Adjust line width
        chart.series[1].graphicalProperties.line.solidFill = "FF0000"  # Red color
        chart.series[1].graphicalProperties.line.width = 20000  # Adjust line width
        chart.series[1].graphicalProperties.line.dashStyle = "dash"

        # Add chart to sheet
        sheet.add_chart(chart, f"E{start_row}")

        # Adjust chart size
        chart.width = 15
        chart.height = 10

    def add_summary_sheet(self):
        sheet = self.workbook.create_sheet("Summary")
        sheet.column_dimensions['A'].width = 50
        sheet.column_dimensions['B'].width = 20
        
        sheet['A1'] = "Database Quality Report Summary"
        sheet['A1'].font = Font(bold=True, size=14)
        sheet.merge_cells('A1:B1')
        
        patient_count = self.all_stats['patient_count'][0][0]

        
        indicators = [
            ("Total Number of Patients", f"{patient_count}"),
            ("Total Number of Documents", "=Document_Counts!B2"),
            ("Number of Document Origins", "=Document_Counts!B3"),
            ("Recent Documents Uploaded (Last 7 Days)", "=Recent_Documents!B2"),
            ("Recent Document Uploaded Sources", "=Recent_Documents!B3")
        ]

        for i, (indicator, formula) in enumerate(indicators, start=3):
            sheet[f'A{i}'] = indicator
            sheet[f'B{i}'] = formula
            sheet[f'B{i}'].number_format = '#,##0'

        self.apply_table_style(sheet, 'A3:B7')

        # Add new section for special patient types
        sheet['A9'] = "Special Patient Types"
        sheet['A9'].font = Font(bold=True, size=12)

        special_patients = [
            ("Special patients", ''),
            ("Test Patients (LASTNAME = 'TEST')", self.all_stats['test_patients']),
            ("Research Patients (LASTNAME = 'INSECTE')", self.all_stats['research_patients']),
            ("Celebrity Patients (LASTNAME = 'FLEUR')", self.all_stats['celebrity_patients'])
        ]

        for i, (patient_type, patients) in enumerate(special_patients, start=10):
            sheet[f'A{i}'] = patient_type
            sheet[f'B{i}'] = len(patients)
            sheet[f'B{i}'].number_format = '#,##0'

        self.apply_table_style(sheet, f'A10:B{13}')

    def add_document_counts_sheet(self):
        sheet = self.workbook.create_sheet("Document_Counts")
        doc_counts = self.all_stats['document_counts']

        # Process and group the data
        grouped_counts = {}
        other_count = 0
        total_count = sum(count for _, count in doc_counts)

        for origin, count in doc_counts:
            percentage = count / total_count
            if origin.startswith("Easily"):
                grouped_counts["Easily"] = grouped_counts.get("Easily", 0) + count
            elif origin.startswith("DOC_EXTERNE"):
                grouped_counts["DOC_EXTERNE"] = grouped_counts.get("DOC_EXTERNE", 0) + count
            elif percentage < 0.01:
                other_count += count
            else:
                grouped_counts[origin] = count

        if other_count > 0:
            grouped_counts["Other"] = other_count

        # Sort the grouped counts by value in descending order
        sorted_counts = sorted(grouped_counts.items(), key=lambda x: x[1], reverse=True)

        sheet['A1'] = "Document Counts by Origin"
        sheet['A1'].font = Font(bold=True, size=14)
        sheet['A2'] = "Total Documents"
        sheet['B2'] = total_count
        sheet['A3'] = "Number of Origins"
        sheet['B3'] = len(sorted_counts)

        sheet['A5'] = "Document Origin Code"
        sheet['B5'] = "Percentage"
        sheet['C5'] = "Unique Document Count"

        for i, (origin, count) in enumerate(sorted_counts, start=6):
            sheet[f'A{i}'] = origin
            sheet[f'B{i}'] = f"=C{i}/SUM($C$6:$C${len(sorted_counts)+5})"
            sheet[f'C{i}'] = count

        self.apply_table_style(sheet, f'A5:C{len(sorted_counts)+5}')
        sheet['B2'].number_format = '#,##0'
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 25

        for row in sheet[f'B6:B{len(sorted_counts)+5}']:
            row[0].number_format = '0.00%'

        for row in sheet[f'C6:C{len(sorted_counts)+5}']:
            row[0].number_format = '#,##0'

        self.add_pie_chart(sheet, f'A5:C{len(sorted_counts)+5}', 'E5', "Document Distribution by Origin")
        
        # Add line graphs for each DOCUMENT_ORIGIN_CODE
        origins = [row[0] for row in self.all_stats['document_origins']]
        for origin in origins:
            self.add_document_count_graph(sheet, origin, by_year=True)

    def add_recent_document_counts_sheet(self):
        sheet = self.workbook.create_sheet("Recent_Documents")
        doc_counts = self.all_stats['recent_document_counts']

        # Process and group the data
        grouped_counts = {}
        total_count = sum(count for _, count in doc_counts)

        for origin, count in doc_counts:
            percentage = count / total_count
            if origin.startswith("Easily"):
                grouped_counts["Easily"] = grouped_counts.get("Easily", 0) + count
            elif origin.startswith("DOC_EXTERNE"):
                grouped_counts["DOC_EXTERNE"] = grouped_counts.get("DOC_EXTERNE", 0) + count
            else:
                grouped_counts[origin] = count
        # Sort the grouped counts by value in descending order
        sorted_counts = sorted(grouped_counts.items(), key=lambda x: x[1], reverse=True)

        sheet['A1'] = "Recent Document Counts by Origin (Last 7 Days)"
        sheet['A1'].font = Font(bold=True, size=14)
        sheet['A2'] = "Total Recent Documents"
        sheet['B2'] = total_count
        sheet['A3'] = "Number of Origins"
        sheet['B3'] = len(sorted_counts)

        sheet['A5'] = "Document Origin Code"
        sheet['B5'] = "Percentage"
        sheet['C5'] = "Unique Document Count"

        for i, (origin, count) in enumerate(sorted_counts, start=6):
            sheet[f'A{i}'] = origin
            sheet[f'B{i}'] = f"=C{i}/SUM($C$13:$C${len(sorted_counts)+5})"
            sheet[f'C{i}'] = count
        origins = [row[0] for row in self.all_stats['document_origins']]
        for origin in origins:
            self.add_document_count_graph(sheet, origin, by_year=False)

        self.apply_table_style(sheet, f'A5:C{len(sorted_counts)+5}')
        sheet['B2'].number_format = '#,##0'
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 25

        for row in sheet[f'B6:B{len(sorted_counts)+5}']:
            row[0].number_format = '0.00%'

        for row in sheet[f'C6:C{len(sorted_counts)+5}']:
            row[0].number_format = '#,##0'

        self.add_pie_chart(sheet, f'A5:C9', '05', "Recent Document Distribution by Origin")

    def add_document_count_graph(self, sheet, origin_code, by_year=True):
        if by_year:
            counts = self.all_stats['document_counts_by_year'].get(origin_code, [])
            title = f"Document Count by Year - {origin_code}"
            x_axis_title = "Year"
        else:
            counts = self.all_stats['recent_document_counts_by_month'].get(origin_code, [])
            title = f"Recent Document Count by Month - {origin_code}"
            x_axis_title = "Month"

        if not counts:
            self.logger.warning(f"No data available for {origin_code}")
            return

        # Add data to sheet
        start_row = sheet.max_row + 2
        sheet.cell(row=start_row, column=1, value=title)
        sheet.cell(row=start_row + 1, column=1, value=x_axis_title)
        sheet.cell(row=start_row + 1, column=2, value="Document Count")

        for i, (date, count) in enumerate(counts, start=start_row + 2):
            sheet.cell(row=i, column=1, value=date)
            sheet.cell(row=i, column=2, value=count)

        # Only create chart if there's more than one data point
        if len(counts) > 1:
            # Create chart
            chart = LineChart()
            chart.title = title
            chart.style = 2
            chart.x_axis.title = x_axis_title
            chart.y_axis.title = "Document Count"

            # Add data series
            data = Reference(sheet, min_col=2, min_row=start_row + 1, max_row=start_row + 1 + len(counts))
            series = chart.series.new()
            series.values = data
            series.title = "Document Count"

            # Set categories
            cats = Reference(sheet, min_col=1, min_row=start_row + 2, max_row=start_row + 1 + len(counts))
            chart.set_categories(cats)

            # Add median line
            median = self.db_checker.get_median_document_count(counts)
            median_series = chart.series.new()
            median_series.values = [median] * len(counts)
            median_series.title = f"Median ({median})"

            # Style the series
            chart.series[0].graphicalProperties.line.width = 20000  # Adjust line width
            chart.series[1].graphicalProperties.line.solidFill = "FF0000"  # Red color
            chart.series[1].graphicalProperties.line.width = 20000  # Adjust line width
            chart.series[1].graphicalProperties.line.dashStyle = "dash"

            # Add chart to sheet, ensuring a valid cell reference
            chart_cell = f"E{start_row}"
            if chart_cell[0].isalpha() and chart_cell[1:].isdigit():
                sheet.add_chart(chart, chart_cell)
            else:
                self.logger.warning(f"Invalid chart position {chart_cell} for {origin_code}. Skipping chart.")

            # Adjust chart size
            chart.width = 15
            chart.height = 10
        else:
            self.logger.warning(f"Not enough data points to create chart for {origin_code}")

    def add_top_users_sheet(self):
        sheet = self.workbook.create_sheet("Top Users")
        top_users = self.all_stats['top_users']
        
        sheet['A1'] = "Top Users by Query Count"
        sheet['A1'].font = Font(bold=True, size=14)
        
        sheet['A3'] = "First Name"
        sheet['B3'] = "Last Name"
        sheet['C3'] = "Query Count"
        
        for i, (firstname, lastname, count) in enumerate(top_users, start=4):
            sheet[f'A{i}'] = firstname
            sheet[f'B{i}'] = lastname
            sheet[f'C{i}'] = count
        
        self.apply_table_style(sheet, f'A3:C{len(top_users)+3}')
        self.add_column_chart(sheet, f'A3:C{len(top_users)+3}', 'E3', "Top Users by Query Count")


    def add_top_users_current_year_sheet(self):
        sheet = self.workbook.create_sheet("Top Users Current Year")
        top_users = self.all_stats['top_users_current_year']
        
        current_year = datetime.now().year
        sheet['A1'] = f"Top Users by Query Count (Current Year: {current_year})"
        sheet['A1'].font = Font(bold=True, size=14)
        
        sheet['A3'] = "First Name"
        sheet['B3'] = "Last Name"
        sheet['C3'] = "Query Count"
        
        for i, (firstname, lastname, count) in enumerate(top_users, start=4):
            sheet[f'A{i}'] = firstname
            sheet[f'B{i}'] = lastname
            sheet[f'C{i}'] = count
        
        self.apply_table_style(sheet, f'A3:C{len(top_users)+3}')
        self.add_column_chart(sheet, f'A3:C{len(top_users)+3}', 'E3', f"Top Users by Query Count (Current Year: {current_year})")


    def add_document_metrics_sheet(self):
        sheet = self.workbook.create_sheet("Data Quality Checks")
        sheet.column_dimensions['A'].width = 40
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        
        sheet['A1'] = "Document Metrics"
        sheet['A1'].font = Font(bold=True, size=14)
        sheet.merge_cells('A1:E1')

        # Fetch statistics and delay data
        stats_list = self.all_stats['stats_and_delays']
        delay_results = self.all_stats['delay_results']
        
        print(f"Debug - stats_list: {stats_list}")
        print(f"Debug - delay_results type: {type(delay_results)}")
        print(f"Debug - delay_results sample: {delay_results[:5] if delay_results else 'None'}")

        if not stats_list or not delay_results:
            sheet['A2'] = "Error: Unable to retrieve document delay data."
            sheet['A2'].font = Font(color="FF0000", bold=True)
            return

        try:
            min_delay, q1, median, q3, max_delay, avg_delay = stats_list[0]
        except (IndexError, ValueError) as e:
            sheet['A2'] = f"Error: Unexpected format in stats_list. {str(e)}"
            sheet['A2'].font = Font(color="FF0000", bold=True)
            return

        # Add statistics to the sheet
        headers = ["Statistic", "Value (days)"]
        
        data = [
            ("Minimum", min_delay),
            ("First Quartile (Q1)", q1),
            ("Median", median),
            ("Third Quartile (Q3)", q3),
            ("Maximum", max_delay),
            ("Average", avg_delay)
        ]
        
        # Add headers
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # Add data
        for row, (label, value) in enumerate(data, start=3):
            sheet.cell(row=row, column=1, value=label)
            cell = sheet.cell(row=row, column=2, value=value)
            cell.number_format = '0.00'
            
            # Add conditional formatting for negative values
            if value < 0:
                cell.font = Font(color="FF0000")  # Red color for negative values
        
        # Add border to the table
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in sheet['A2:B8']:
            for cell in row:
                cell.border = thin_border
        
        # Add a note about negative values
        note = sheet.cell(row=9, column=1, value="Note: Negative values indicate documents updated before their creation date in the DPI. We filtered out the Doctolib documents for these metrics.")
        note.font = Font(italic=True)
        
        # Add min and max delay document details
        if delay_results and isinstance(delay_results, list) and len(delay_results) > 0:
            sheet['A11'] = "Minimum and Maximum Delay Documents"
            sheet['A11'].font = Font(bold=True)

            headers = ["Delay Type", "Title", "Document Creation Date", "Document Origin", "Upload EDS Date", "Delay (Days)"]
            for col, header in enumerate(headers, start=1):
                sheet.cell(row=13, column=col, value=header).font = Font(bold=True)

            for row, doc in enumerate(delay_results, start=14):
                if len(doc) >= 6:
                    sheet.cell(row=row, column=1, value=doc[5])  # Delay Type
                    sheet.cell(row=row, column=2, value=doc[0])  # Title
                    sheet.cell(row=row, column=3, value=doc[1])  # Document Date
                    sheet.cell(row=row, column=4, value=doc[2])  # Document Origin Code
                    sheet.cell(row=row, column=5, value=doc[3])  # Update Date
                    sheet.cell(row=row, column=6, value=doc[4])  # Delay Days
                    sheet.cell(row=row, column=6).number_format = '0.00'
                else:
                    sheet.cell(row=row, column=1, value=f"Error: Invalid data format for row {row}")
        else:
            sheet['A11'] = "Unable to retrieve minimum and maximum delay document details."
            sheet['A11'].font = Font(color="FF0000", italic=True)

        print("Debug - Finished add_document_metrics_sheet method")

    def apply_table_style(self, sheet, cell_range):
        light_blue_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for row in sheet[cell_range]:
            for cell in row:
                cell.border = border
                if cell.row == sheet[cell_range.split(':')[0]].row:
                    cell.font = Font(bold=True)
                    cell.fill = light_blue_fill
                cell.alignment = Alignment(horizontal='left')

    def add_pie_chart(self, sheet, data_range, position, title):
        pie = PieChart()
        labels = Reference(sheet, min_col=1, min_row=6, max_row=sheet.max_row)
        data = Reference(sheet, min_col=3, min_row=5, max_row=sheet.max_row)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = title


        # Keep legend
        pie.legend.position = 'r'  # Position legend to the right

        # Optional: Customize the chart layout
        pie.height = 15  # Set the height of the chart
        pie.width = 20   # Set the width of the chart

        sheet.add_chart(pie, position)

    def add_line_chart(self, sheet, data_range, position, title):
        chart = LineChart()
        chart.style = 10
        chart.title = title
        chart.y_axis.title = 'Query Count'
        chart.x_axis.title = 'User'

        data = Reference(sheet, min_col=3, min_row=3, max_row=sheet[data_range.split(':')[1]].row)
        cats = Reference(sheet, min_col=1, min_row=4, max_row=sheet[data_range.split(':')[1]].row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Remove all ticks and gridlines
        #chart.y_axis.majorTickMark = "none"
        #chart.y_axis.minorTickMark = "none"
        #chart.x_axis.majorTickMark = "none"
        #chart.x_axis.minorTickMark = "none"

        sheet.add_chart(chart, position)

    def add_column_chart(self, sheet, data_range, position, title):
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = title
        chart.y_axis.title = 'Query Count'
        chart.x_axis.title = 'User'

        data = Reference(sheet, min_col=3, min_row=3, max_row=sheet[data_range.split(':')[1]].row)
        cats = Reference(sheet, min_col=1, min_row=4, max_row=sheet[data_range.split(':')[1]].row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Remove all ticks 
        chart.y_axis.majorTickMark = "none"
        chart.y_axis.minorTickMark = "none"
        chart.x_axis.majorTickMark = "none"
        chart.x_axis.minorTickMark = "none"

        sheet.add_chart(chart, position)


if __name__ == "__main__":
    configure_logging()
    logger = logging.getLogger(__name__)
    logger.info("Starting database quality report generation")
    
    report_generator = DatabaseQualityReportGenerator()
    report_generator.generate_report()
    
    logger.info("Database quality report generation completed")