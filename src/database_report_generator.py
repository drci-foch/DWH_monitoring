import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.axis import ChartLines
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from queries import DatabaseQualityChecker
from datetime import datetime
import numpy as np

class DatabaseQualityReportGenerator:
    def __init__(self):
        self.db_checker = DatabaseQualityChecker()
        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet

    def generate_report(self, filename=r"P:\RechercheClinique\Unité DATA\EDS Foch\Documentation EDS\Documentation technique\Documentation_base_DWH\Monitoring DWH\database_quality_report_{}.xlsx"):
        current_date = datetime.now().strftime("%d-%m-%Y")
        filename = filename.format(current_date)        
        self.add_summary_sheet()
        self.add_document_metrics_sheet()
        self.add_document_counts_sheet()
        self.add_recent_document_counts_sheet()
        self.add_top_users_sheet()
        self.add_top_users_current_year_sheet()
        self.workbook.save(filename)
        print(f"Report generated: {filename}")

    def add_summary_sheet(self):
        sheet = self.workbook.create_sheet("Summary")
        sheet.column_dimensions['A'].width = 50
        sheet.column_dimensions['B'].width = 20
        
        sheet['A1'] = "Database Quality Report Summary"
        sheet['A1'].font = Font(bold=True, size=14)
        sheet.merge_cells('A1:B1')
        
        patient_count = self.db_checker.get_patient_count()
        
        indicators = [
            ("Total Number of Patients", f"{patient_count}"),
            ("Total Number of Documents", "=Document_Counts!B2"),
            ("Number of Document Origins", "=Document_Counts!B3"),
            ("Recent Documents Uploaded (Last 7 Days)", "=Recent_Documents!B2"),
            ("Recent Document Uploaded Sources", "=Recent_Documents!B3")
        ]

        for i, (indicator, formula) in enumerate(indicators, start=3):
            sheet[f'A{i}'] = indicator
            sheet[f'B{i}'] = formula
            sheet[f'B{i}'].number_format = '#,##0'

        self.apply_table_style(sheet, 'A3:B7')

    # def add_patient_count_sheet(self):
    #     sheet = self.workbook.create_sheet("Patient_Count")
    #     patient_count = self.db_checker.get_patient_count()

    #     sheet['A1'] = "Total Number of Patients"
    #     sheet['A1'].font = Font(bold=True, size=14)
    #     sheet['A2'] = "Count"
    #     sheet['B2'] = patient_count
    #     sheet['B2'].number_format = '#,##0'

    #     self.apply_table_style(sheet, 'A2:B2')

    def add_document_counts_sheet(self):
        sheet = self.workbook.create_sheet("Document_Counts")
        doc_counts = self.db_checker.get_document_counts_by_origin()

        # Process and group the data
        grouped_counts = {}
        other_count = 0
        total_count = sum(count for _, count in doc_counts)

        for origin, count in doc_counts:
            percentage = count / total_count
            if origin.startswith("Easily"):
                grouped_counts["Easily"] = grouped_counts.get("Easily", 0) + count
            elif origin.startswith("DOC_EXTERNE"):
                grouped_counts["DOC_EXTERNE"] = grouped_counts.get("DOC_EXTERNE", 0) + count
            elif percentage < 0.01:
                other_count += count
            else:
                grouped_counts[origin] = count

        if other_count > 0:
            grouped_counts["Other"] = other_count

        # Sort the grouped counts by value in descending order
        sorted_counts = sorted(grouped_counts.items(), key=lambda x: x[1], reverse=True)

        sheet['A1'] = "Document Counts by Origin"
        sheet['A1'].font = Font(bold=True, size=14)
        sheet['A2'] = "Total Documents"
        sheet['B2'] = total_count
        sheet['A3'] = "Number of Origins"
        sheet['B3'] = len(sorted_counts)

        sheet['A5'] = "Document Origin Code"
        sheet['B5'] = "Percentage"
        sheet['C5'] = "Unique Document Count"

        for i, (origin, count) in enumerate(sorted_counts, start=6):
            sheet[f'A{i}'] = origin
            sheet[f'B{i}'] = f"=C{i}/SUM($C$6:$C${len(sorted_counts)+5})"
            sheet[f'C{i}'] = count

        self.apply_table_style(sheet, f'A5:C{len(sorted_counts)+5}')
        sheet['B2'].number_format = '#,##0'
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 25

        for row in sheet[f'B6:B{len(sorted_counts)+5}']:
            row[0].number_format = '0.00%'

        for row in sheet[f'C6:C{len(sorted_counts)+5}']:
            row[0].number_format = '#,##0'

        self.add_pie_chart(sheet, f'A5:C{len(sorted_counts)+5}', 'E5', "Document Distribution by Origin")

    def add_recent_document_counts_sheet(self):
        sheet = self.workbook.create_sheet("Recent_Documents")
        doc_counts = self.db_checker.get_recent_document_counts_by_origin()

        # Process and group the data
        grouped_counts = {}
        total_count = sum(count for _, count in doc_counts)

        for origin, count in doc_counts:
            percentage = count / total_count
            if origin.startswith("Easily"):
                grouped_counts["Easily"] = grouped_counts.get("Easily", 0) + count
            elif origin.startswith("DOC_EXTERNE"):
                grouped_counts["DOC_EXTERNE"] = grouped_counts.get("DOC_EXTERNE", 0) + count
            else:
                grouped_counts[origin] = count
        # Sort the grouped counts by value in descending order
        sorted_counts = sorted(grouped_counts.items(), key=lambda x: x[1], reverse=True)

        sheet['A1'] = "Recent Document Counts by Origin (Last 7 Days)"
        sheet['A1'].font = Font(bold=True, size=14)
        sheet['A2'] = "Total Recent Documents"
        sheet['B2'] = total_count
        sheet['A3'] = "Number of Origins"
        sheet['B3'] = len(sorted_counts)

        sheet['A5'] = "Document Origin Code"
        sheet['B5'] = "Percentage"
        sheet['C5'] = "Unique Document Count"

        for i, (origin, count) in enumerate(sorted_counts, start=6):
            sheet[f'A{i}'] = origin
            sheet[f'B{i}'] = f"=C{i}/SUM($C$6:$C${len(sorted_counts)+5})"
            sheet[f'C{i}'] = count

        self.apply_table_style(sheet, f'A5:C{len(sorted_counts)+5}')
        sheet['B2'].number_format = '#,##0'
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 25

        for row in sheet[f'B6:B{len(sorted_counts)+5}']:
            row[0].number_format = '0.00%'

        for row in sheet[f'C6:C{len(sorted_counts)+5}']:
            row[0].number_format = '#,##0'

        self.add_pie_chart(sheet, f'A5:C{len(sorted_counts)+5}', 'E5', "Recent Document Distribution by Origin")


    def add_top_users_sheet(self):
        sheet = self.workbook.create_sheet("Top Users")
        top_users = self.db_checker.get_top_users()
        
        sheet['A1'] = "Top Users by Query Count"
        sheet['A1'].font = Font(bold=True, size=14)
        
        sheet['A3'] = "First Name"
        sheet['B3'] = "Last Name"
        sheet['C3'] = "Query Count"
        
        for i, (firstname, lastname, count) in enumerate(top_users, start=4):
            sheet[f'A{i}'] = firstname
            sheet[f'B{i}'] = lastname
            sheet[f'C{i}'] = count
        
        self.apply_table_style(sheet, f'A3:C{len(top_users)+3}')
        self.add_column_chart(sheet, f'A3:C{len(top_users)+3}', 'E3', "Top Users by Query Count")


    def add_top_users_current_year_sheet(self):
        sheet = self.workbook.create_sheet("Top Users Current Year")
        top_users = self.db_checker.get_top_users_current_year()
        
        current_year = datetime.now().year
        sheet['A1'] = f"Top Users by Query Count (Current Year: {current_year})"
        sheet['A1'].font = Font(bold=True, size=14)
        
        sheet['A3'] = "First Name"
        sheet['B3'] = "Last Name"
        sheet['C3'] = "Query Count"
        
        for i, (firstname, lastname, count) in enumerate(top_users, start=4):
            sheet[f'A{i}'] = firstname
            sheet[f'B{i}'] = lastname
            sheet[f'C{i}'] = count
        
        self.apply_table_style(sheet, f'A3:C{len(top_users)+3}')
        self.add_column_chart(sheet, f'A3:C{len(top_users)+3}', 'E3', f"Top Users by Query Count (Current Year: {current_year})")


    def add_document_metrics_sheet(self):
        sheet = self.workbook.create_sheet("Data Quality Checks")
        sheet.column_dimensions['A'].width = 40
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        
        sheet['A1'] = "Document Metrics"
        sheet['A1'].font = Font(bold=True, size=14)
        sheet.merge_cells('A1:E1')

        # Fetch statistics and delay data
        stats_list, delay_results = self.db_checker.get_stats_and_delays_document_monthly()
        
        print(f"Debug - stats_list: {stats_list}")
        print(f"Debug - delay_results type: {type(delay_results)}")
        print(f"Debug - delay_results sample: {delay_results[:5] if delay_results else 'None'}")

        if not stats_list or not delay_results:
            sheet['A2'] = "Error: Unable to retrieve document delay data."
            sheet['A2'].font = Font(color="FF0000", bold=True)
            return

        try:
            min_delay, q1, median, q3, max_delay, avg_delay = stats_list[0]
        except (IndexError, ValueError) as e:
            sheet['A2'] = f"Error: Unexpected format in stats_list. {str(e)}"
            sheet['A2'].font = Font(color="FF0000", bold=True)
            return

        # Add statistics to the sheet
        headers = ["Statistic", "Value (days)"]
        
        data = [
            ("Minimum", min_delay),
            ("First Quartile (Q1)", q1),
            ("Median", median),
            ("Third Quartile (Q3)", q3),
            ("Maximum", max_delay),
            ("Average", avg_delay)
        ]
        
        # Add headers
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # Add data
        for row, (label, value) in enumerate(data, start=3):
            sheet.cell(row=row, column=1, value=label)
            cell = sheet.cell(row=row, column=2, value=value)
            cell.number_format = '0.00'
            
            # Add conditional formatting for negative values
            if value < 0:
                cell.font = Font(color="FF0000")  # Red color for negative values
        
        # Add border to the table
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in sheet['A2:B8']:
            for cell in row:
                cell.border = thin_border

        # # Extract delay days for histogram
        # try:
        #     all_delays = [float(row[4]) for row in delay_results if row and len(row) > 4]
        # except (IndexError, ValueError) as e:
        #     sheet['A9'] = f"Error extracting delay data: {str(e)}"
        #     sheet['A9'].font = Font(color="FF0000", bold=True)
        #     return

        # if not all_delays:
        #     sheet['A9'] = "No valid delay data found for histogram."
        #     sheet['A9'].font = Font(color="FF0000", bold=True)
        #     return

        # # Create histogram data
        # hist, bin_edges = np.histogram(all_delays, bins=20)
        
        # # Prepare data for histogram
        # sheet['D2'] = "Histogram Data"
        # sheet['D2'].font = Font(bold=True)
        # sheet['E2'] = "Frequency"
        # sheet['E2'].font = Font(bold=True)
        
        # for i, (edge, freq) in enumerate(zip(bin_edges[:-1], hist), start=3):
        #     sheet.cell(row=i, column=4, value=f"{edge:.2f} to {bin_edges[i-2]:.2f}")
        #     sheet.cell(row=i, column=5, value=freq)
        
        # # Create bar chart (histogram)
        # chart = BarChart()
        # chart.type = "col"
        # chart.style = 10
        # chart.title = "Document Delay Distribution (past month)"
        # chart.y_axis.title = "Frequency"
        # chart.x_axis.title = "Delay Range (days)"
        
        # data = Reference(sheet, min_col=5, min_row=2, max_row=22, max_col=5)
        # cats = Reference(sheet, min_col=4, min_row=3, max_row=22)
        # chart.add_data(data, titles_from_data=True)
        # chart.set_categories(cats)
        
        # # Customize chart
        # chart.series[0].graphicalProperties.solidFill = "8BB9E7"
        # chart.y_axis.majorGridlines = None
        # chart.x_axis.tickLblPos = "low"
        
        # sheet.add_chart(chart, "A10")
        
        # # Adjust chart size
        # chart.width = 30
        # chart.height = 15

        # sheet.sheet_view.showGridLines = False
        
        # Add a note about negative values
        note = sheet.cell(row=9, column=1, value="Note: Negative values indicate documents updated before their creation date in the DPI. We filtered out the Doctolib documents for these metrics.")
        note.font = Font(italic=True)
        
        # Add min and max delay document details
        if delay_results and isinstance(delay_results, list) and len(delay_results) > 0:
            sheet['A11'] = "Minimum and Maximum Delay Documents"
            sheet['A11'].font = Font(bold=True)

            headers = ["Delay Type", "Title", "Document Creation Date", "Document Origin", "Upload EDS Date", "Delay (Days)"]
            for col, header in enumerate(headers, start=1):
                sheet.cell(row=13, column=col, value=header).font = Font(bold=True)

            for row, doc in enumerate(delay_results, start=14):
                if len(doc) >= 6:
                    sheet.cell(row=row, column=1, value=doc[5])  # Delay Type
                    sheet.cell(row=row, column=2, value=doc[0])  # Title
                    sheet.cell(row=row, column=3, value=doc[1])  # Document Date
                    sheet.cell(row=row, column=4, value=doc[2])  # Document Origin Code
                    sheet.cell(row=row, column=5, value=doc[3])  # Update Date
                    sheet.cell(row=row, column=6, value=doc[4])  # Delay Days
                    sheet.cell(row=row, column=6).number_format = '0.00'
                else:
                    sheet.cell(row=row, column=1, value=f"Error: Invalid data format for row {row}")
        else:
            sheet['A11'] = "Unable to retrieve minimum and maximum delay document details."
            sheet['A11'].font = Font(color="FF0000", italic=True)

        print("Debug - Finished add_document_metrics_sheet method")

    def apply_table_style(self, sheet, cell_range):
        light_blue_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for row in sheet[cell_range]:
            for cell in row:
                cell.border = border
                if cell.row == sheet[cell_range.split(':')[0]].row:
                    cell.font = Font(bold=True)
                    cell.fill = light_blue_fill

    def add_pie_chart(self, sheet, data_range, position, title):
        pie = PieChart()
        labels = Reference(sheet, min_col=1, min_row=6, max_row=sheet.max_row)
        data = Reference(sheet, min_col=3, min_row=5, max_row=sheet.max_row)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = title

        # pie.dataLabels = DataLabelList()
        # pie.dataLabels.showCatName = False
        # pie.dataLabels.showVal = False
        # pie.dataLabels.showPercent = False

        # Keep legend
        pie.legend.position = 'r'  # Position legend to the right

        # Optional: Customize the chart layout
        pie.height = 15  # Set the height of the chart
        pie.width = 20   # Set the width of the chart

        sheet.add_chart(pie, position)

    def add_line_chart(self, sheet, data_range, position, title):
        chart = LineChart()
        chart.style = 10
        chart.title = title
        chart.y_axis.title = 'Query Count'
        chart.x_axis.title = 'User'

        data = Reference(sheet, min_col=3, min_row=3, max_row=sheet[data_range.split(':')[1]].row)
        cats = Reference(sheet, min_col=1, min_row=4, max_row=sheet[data_range.split(':')[1]].row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Remove all ticks and gridlines
        #chart.y_axis.majorGridlines = None
        #chart.y_axis.minorGridlines = None
        #chart.x_axis.majorGridlines = None
        #chart.x_axis.minorGridlines = None
        chart.y_axis.majorTickMark = "none"
        chart.y_axis.minorTickMark = "none"
        chart.x_axis.majorTickMark = "none"
        chart.x_axis.minorTickMark = "none"
        #chart.y_axis.delete = True  # This removes the axis line
        #chart.x_axis.delete = True  # This removes the axis line

        sheet.add_chart(chart, position)

    def add_column_chart(self, sheet, data_range, position, title):
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = title
        chart.y_axis.title = 'Query Count'
        chart.x_axis.title = 'User'

        data = Reference(sheet, min_col=3, min_row=3, max_row=sheet[data_range.split(':')[1]].row)
        cats = Reference(sheet, min_col=1, min_row=4, max_row=sheet[data_range.split(':')[1]].row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Remove all ticks and gridlines
        #chart.y_axis.majorGridlines = None
        #chart.y_axis.minorGridlines = None
        #chart.x_axis.majorGridlines = None
        #chart.x_axis.minorGridlines = None
        chart.y_axis.majorTickMark = "none"
        chart.y_axis.minorTickMark = "none"
        chart.x_axis.majorTickMark = "none"
        chart.x_axis.minorTickMark = "none"
        #chart.y_axis.delete = True  # This removes the axis line
        #chart.x_axis.delete = True  # This removes the axis line

        sheet.add_chart(chart, position)


if __name__ == "__main__":
    report_generator = DatabaseQualityReportGenerator()
    report_generator.generate_report()